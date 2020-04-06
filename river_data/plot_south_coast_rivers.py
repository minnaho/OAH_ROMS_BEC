#############################################
# plot_south_coast_rivers.py
# plot data from Final River Compilation and
# rational methods file (SCB_RIVERS.mat)  
# from netcdf files 
# south_coast_rivers_10_years_no_watershed.nc
# south_coast_rivers_24_years.nc
#####################################################
import numpy as np
import pickle
import matplotlib.pyplot as plt
import matplotlib.dates as mdate 
import datetime
import copy
from collections import defaultdict
from netCDF4 import Dataset, date2num, num2date
import openpyxl
from openpyxl import Workbook

###################
# FOLDER PATHS
####################
grid_path = '../usw1_grd.nc'
save_figs_path = './river_figs/'

################################
# CALL DATA FROM netcdf files  
################################

# get river names
names_10 = pickle.load(open('river_names_10.pkl','rb'))
names_24 = pickle.load(open('river_names_24.pkl','rb'))

# load river netcdf files
nc_10 = Dataset('south_coast_rivers_10_years_no_watershed_new.nc','r')
nc_24 = Dataset('south_coast_rivers_24_years_new.nc','r')

time_10 = nc_10.variables['time']
lats_10 = nc_10.variables['latitude']
lons_10 = nc_10.variables['longitude']
flow_10 = nc_10.variables['flow']
NH4_10 = nc_10.variables['ammonium']
NO3_10 = nc_10.variables['nitrate']
PO4_10 = nc_10.variables['phosphate']
TN_10 = nc_10.variables['total_nitrogen']
TP_10 = nc_10.variables['total_phosphorus']

date_10 = num2date(np.asarray(time_10),time_10.units)

time_24 = nc_24.variables['time']
lats_24 = nc_24.variables['latitude']
lons_24 = nc_24.variables['longitude']
flow_24 = nc_24.variables['flow']
NH4_24 = nc_24.variables['ammonium']
NO3_24 = nc_24.variables['nitrate']
PO4_24 = nc_24.variables['phosphate']
TN_24 = nc_24.variables['total_nitrogen']
TP_24 = nc_24.variables['total_phosphorus']

date_24 = num2date(np.asarray(time_24),time_24.units)

lats_all = np.concatenate((lats_10,lats_24))
lons_all = np.concatenate((lons_10,lons_24))
river_lat_lon = np.array((lats_all,lons_all))
np.save('river_lat_lon',river_lat_lon)

v = [
     'time',
     'flow',
     'ammonium',
     'nitrate',
     'phosphate',
     'total_nitrogen',
     'total_phosphorus']

nutrients = ['NH4','NO3','PO4']

####################################################################
# get total nitrogen and phosphorus loads per year in kg
# multiply TN and TP by flow (m3/s)*(s/day)*(mmol/m3)*(mg/mmol)*(kg/mg)
# sum up the loads for each day over a year
#####################################################################
mol_wt_N = 14.007 # total nitrogen
mol_wt_P = 30.974 # total phosphorus
seconds_in_day = 86400
mg_to_kg = 1./1e6
'''
# create array of daily loads
load_TN_day_10 = np.zeros((len(time_10),len(TN_10[0])))
load_TP_day_10 = np.zeros((len(time_10),len(TP_10[0])))

load_TN_day_24 = np.zeros((len(time_24),len(TN_24[0])))
load_TP_day_24 = np.zeros((len(time_24),len(TP_24[0])))

# calculate daily load
for t_c in range(len(time_10)):
    for r_c in range(len(flow_10[1])):
        load_TN_day_10[t_c,r_c] = TN_10[t_c,r_c,r_c] * flow_10[t_c,r_c,r_c] * seconds_in_day * mol_wt_N * mg_to_kg
        load_TP_day_10[t_c,r_c] = TP_10[t_c,r_c,r_c] * flow_10[t_c,r_c,r_c] * seconds_in_day * mol_wt_P * mg_to_kg
    print('calculating daily load for compilation data time '+str(t_c)+' of '+str(len(time_10)))


for t_r in range(len(time_24)):
    for r_r in range(len(flow_24[1])):
        load_TN_day_24[t_r,r_r] = TN_24[t_r,r_r,r_r] * flow_24[t_r,r_r,r_r] * seconds_in_day * mol_wt_N * mg_to_kg
        load_TP_day_24[t_r,r_r] = TP_24[t_r,r_r,r_r] * flow_24[t_r,r_r,r_r] * seconds_in_day * mol_wt_P * mg_to_kg
    print('calculating daily load for rational data time '+str(t_r)+' of '+str(len(time_24)))

# create array for yearly loads
# shape (year, river) 
n_years_10 = 10
n_years_24 = 24

load_TN_10 = np.zeros((n_years_10,len(TN_10[0])))
load_TP_10 = np.zeros((n_years_10,len(TP_10[0])))

load_TN_24 = np.zeros((n_years_24,len(TN_24[0])))
load_TP_24 = np.zeros((n_years_24,len(TP_24[0])))

for year in list(range(n_years_10)):
    for river in range(len(TN_10[0])): 
        load_TN_10[year,river] = np.nansum(load_TN_day_10[365*year:365*(year+1)],axis=0)[river]
        load_TP_10[year,river] = np.nansum(load_TP_day_10[365*year:365*(year+1)],axis=0)[river]
    print('calculating sum of compilation data time '+str(year)+' of '+str(n_years_10))
   

for year in list(range(n_years_24)):
    for river in range(len(TN_24[0])):
        load_TN_24[year,river] = np.nansum(load_TN_day_24[365*year:365*(year+1)],axis=0)[river]
        load_TP_24[year,river] = np.nansum(load_TP_day_24[365*year:365*(year+1)],axis=0)[river]
    print('calculating sum of rational data time '+str(year)+' of '+str(n_years_24))



######################
# write to excel file
######################
dest_filename = 'SCB River TN and TP loads updated.xlsx'
wb = openpyxl.load_workbook(dest_filename)
sh = wb.sheetnames

# write nitrogen total loads
row_N = 3 # row/column data starts in in excel
column_c = 2
for y_c in list(range(n_years_10)):
    for r_c in list(range(len(TN_10[0]))):
        wb[sh[0]].cell(row=r_c+row_N,column=y_c+column_c).value = load_TN_10[y_c,r_c]
        wb[sh[1]].cell(row=r_c+row_N,column=y_c+column_c).value = load_TP_10[y_c,r_c]


column_r = 14
for y_r in list(range(n_years_24)):
    for r_r in list(range(len(TN_24[0]))):
        wb[sh[0]].cell(row=r_r+row_N,column=y_r+column_r).value = load_TN_24[y_r,r_r]
        wb[sh[1]].cell(row=r_r+row_N,column=y_r+column_r).value = load_TP_24[y_r,r_r]

wb.save(filename = dest_filename)
'''
'''
###############################
# PLOT NUTRIENT AND FLOW DATA
###############################
colors = ['blue','red','green','purple']
xlabel_size = 14
ylabel_size = 14
title_size = 14
# date indexes for date_24: 1997-1-1 to 2013-12-31
start_ind = 2557
end_ind = 8766

# make one plot for each river with subplots for each variable    
# compilation data (10 years)
for r_i,r_c in enumerate(names_10): 
    plt.figure(figsize=[13,9])
    for v_i in list(range(1,len(v)-2)): 
        plt.subplot(2,2,v_i)
        plt.xlabel('Time',fontsize=xlabel_size)
        if v_i != 1:
            plt.ylabel(v[v_i]+' (mmol/m$^3$)',fontsize=ylabel_size)
        else:
            plt.ylabel(v[v_i]+' (m$^3$/s)',fontsize=ylabel_size) 
        plt.plot(date_10,nc_10.variables[v[v_i]][:,r_i,r_i],linewidth=3,color=colors[v_i-1])
        plt.title(v[v_i],fontsize=title_size)
        locator = mdate.YearLocator()
        plt.gca().xaxis.set_major_locator(locator)
        plt.gcf().autofmt_xdate()
        ax = plt.gca()
        plt.xlim([min(date_10),max(date_10)]) 
        if v_i == 1 or v_i == 2:
            ax.xaxis.set_ticklabels([])
        ax.grid(True)
    plt.savefig(save_figs_path+r_c,bboxinches='tight')
    plt.close('all')
    print(r_c+' plotted')

for r_i,r_r in enumerate(names_24): 
    plt.figure(figsize=[13,9])
    for v_i in list(range(1,len(v)-2)): 
        plt.subplot(2,2,v_i)
        plt.xlabel('Time',fontsize=xlabel_size)
        if v_i != 1:
            plt.ylabel(v[v_i]+' (mmol/m$^3$)',fontsize=ylabel_size)
        else:
            plt.ylabel(v[v_i]+' (m$^3$/s)',fontsize=ylabel_size) 
        plt.plot(date_24,nc_24.variables[v[v_i]][:,r_i,r_i],linewidth=3,color=colors[v_i-1])
        plt.title(v[v_i],fontsize=title_size)
        locator = mdate.YearLocator()
        plt.gca().xaxis.set_major_locator(locator)
        plt.gcf().autofmt_xdate()
        ax = plt.gca() 
        plt.xlim([min(date_24),max(date_24)]) 
        if v_i == 1 or v_i == 2:
            ax.xaxis.set_ticklabels([])
        ax.grid(True)
    plt.savefig(save_figs_path+r_r,bboxinches='tight')
    plt.close('all')
    print(r_r+' plotted')
'''
#########################
# PLOT DAILY LOAD
#########################
months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sept','Oct','Nov','Dec']
colors = ['blue','red','green','purple']
xlabel_size = 14
ylabel_size = 12
title_size = 16
# date indexes for date_24: 1997-1-1 to 2013-12-31
start_ind = 2557
end_ind = 8766
fig_w = 10
fig_h = 8
tick_label_size = 12
adjust_top = .95
hspace = 0

# make one plot for each river with subplots for each variable    
# compilation data (10 years)
for r_i,r_c in enumerate(names_10): 
    fig = plt.figure(figsize=[fig_w,fig_h])
    for v_i in list(range(1,len(v)-2)):  
        # multiply by molecular weight of N or P depending on nutrient
        # when converting to load
        if v_i == 2 or v_i == 3:
            ax = fig.add_subplot(4,1,v_i,sharex=ax)
            ax.set_ylabel(nutrients[v_i-2]+' (kg day$^{-1}$)',fontsize=ylabel_size)
            y = nc_10.variables[v[v_i]][:,r_i,r_i] * flow_10[:,r_i,r_i] * seconds_in_day * mol_wt_N * mg_to_kg
        if v_i == 4:
            ax = fig.add_subplot(4,1,v_i,sharex=ax)           
            ax.set_ylabel(nutrients[v_i-2]+' (kg day$^{-1}$)',fontsize=ylabel_size)
            y = nc_10.variables[v[v_i]][:,r_i,r_i] * flow_10[:,r_i,r_i] * seconds_in_day * mol_wt_P * mg_to_kg

        # ensure correct units in y axis
        if v_i == 1:
            ax = fig.add_subplot(4,1,v_i)
            ax.set_ylabel(v[v_i]+' (m$^3$ s$^{-1}$)',fontsize=ylabel_size) 
            y = nc_10.variables[v[v_i]][:,r_i,r_i]
        plt.plot(date_10[0:365],y[0:365],linewidth=3,color=colors[v_i-1])
        locator = mdate.MonthLocator()
        #plt.gcf().autofmt_xdate()
        a = plt.gca()
        a.xaxis.set_major_locator(locator)
        a.axes.xaxis.set_ticklabels(months)
        ax.tick_params(axis='both',which='major',labelsize=tick_label_size)
        #plt.xlim([min(date_10),max(date_10)]) 
        if v_i != 4:
            a.xaxis.set_ticklabels([])
        a.grid(True)
    plt.suptitle(r_c,fontsize=title_size)
    plt.subplots_adjust(top=adjust_top)
    fig.subplots_adjust(hspace=hspace)
    plt.savefig(save_figs_path+'ts_'+r_c,bboxinches='tight')
    plt.close('all')
    print(r_c+' plotted')
'''
for r_i,r_c in enumerate(names_24): 
    fig = plt.figure(figsize=[fig_w,fig_h])
    for v_i in list(range(1,len(v)-2)):  
        # multiply by molecular weight of N or P depending on nutrient
        # when converting to load
        if v_i == 2 or v_i == 3:
            ax = fig.add_subplot(4,1,v_i,sharex=ax)
            ax.set_ylabel(nutrients[v_i-2]+' (kg day$^{-1}$)',fontsize=ylabel_size)
            y = nc_24.variables[v[v_i]][:,r_i,r_i] * flow_24[:,r_i,r_i] * seconds_in_day * mol_wt_N * mg_to_kg
        if v_i == 4:
            ax = fig.add_subplot(4,1,v_i,sharex=ax)           
            ax.set_ylabel(nutrients[v_i-2]+' (kg day$^{-1}$)',fontsize=ylabel_size)
            y = nc_24.variables[v[v_i]][:,r_i,r_i] * flow_24[:,r_i,r_i] * seconds_in_day * mol_wt_P * mg_to_kg

        # ensure correct units in y axis
        if v_i == 1:
            ax = fig.add_subplot(4,1,v_i)
            ax.set_ylabel(v[v_i]+' (m$^3$ s$^{-1}$)',fontsize=ylabel_size) 
            y = nc_24.variables[v[v_i]][:,r_i,r_i]
        plt.plot(date_24[start_ind:end_ind],y[start_ind:end_ind],linewidth=3,color=colors[v_i-1])
        locator = mdate.YearLocator()
        #plt.gcf().autofmt_xdate()
        a = plt.gca()
        a.xaxis.set_major_locator(locator)
        ax.tick_params(axis='both',which='major',labelsize=tick_label_size)
        #plt.xlim([min(date_10),max(date_10)]) 
        if v_i != 4:
            a.xaxis.set_ticklabels([])
        a.grid(True)
    plt.suptitle(r_c,fontsize=title_size)
    plt.subplots_adjust(top=adjust_top)
    fig.subplots_adjust(hspace=hspace)
    plt.savefig(save_figs_path+'ts_'+r_c,bboxinches='tight')
    plt.close('all')
    print(r_c+' plotted')
'''      
'''
############################
# PLOT RIVER LAT/LON ON GRID
############################
def plot_rivers(x_coords, y_coords, msize):
    [plt.plot(x_coords[n], y_coords[n],'o',markeredgecolor='teal',mfc='teal',markersize=msize) for n in range(len(x_coords))] 


lwidth_shore = 2

# convert lat/lon coordinates to grid coordinates
river_coord_x = []
river_coord_y = []
for coord in range(len(lats_10)):
    lat_you_want = lats_10[coord]
    lon_you_want = lons_10[coord]
    temp = np.abs( (lat_nc - lat_you_want)**2 + (lon_nc - lon_you_want)**2)
    eta_coord,xi_coord = np.unravel_index(temp.argmin(),temp.shape)
    river_coord_x.append(xi_coord)
    river_coord_y.append(eta_coord)

for coord in range(len(lats_24)):
    lat_you_want = lats_24[coord]
    lon_you_want = lons_24[coord]
    temp = np.abs( (lat_nc - lat_you_want)**2 + (lon_nc - lon_you_want)**2)
    eta_coord,xi_coord = np.unravel_index(temp.argmin(),temp.shape)
    river_coord_x.append(xi_coord)
    river_coord_y.append(eta_coord)
'''
'''
# move river locations that are in ocean to 
# first grid cell on land 
# (next for loop ensures it is on first cell off land in ocean)
for i in range(len(river_coord_x)):
    while mask[river_coord_y[i],river_coord_x[i]] == 1:
        river_coord_x[i] += 1     
     
# move river locations that are on land offshore in x direction
# to first grid cell in ocean
for i in range(len(river_coord_x)):
    while mask[river_coord_y[i],river_coord_x[i]] == 0:
        river_coord_x[i] -= 1 
'''
'''
# check if any river locations are still on land 
n = 0
for i in range(len(river_coord_x)):
    if mask[river_coord_y[i],river_coord_x[i]] == 0:
        n+=1
print('number of river positions on land: '+str(n))

#river_x = [i * pm_mult for i in river_coord_x]
#river_y = [i * pn_mult for i in river_coord_y]
river_x = river_coord_x 
river_y = river_coord_y 
river_coords = np.array((river_coord_y,river_coord_x))
np.save('river_coords.npy',river_coords)

plt.figure(figsize=[9,14])
#im = plt.imshow(mask,origin='lower',extent=im_ext_km)
plot_rivers(river_x,river_y,3)
im = plt.imshow(mask,origin='lower')
#plt.contour(h,[5],colors='k',linewidths=lwidth_shore)
plt.xlim([650,Lx])
plt.ylim([125,550])
plt.savefig('river_locs.png',bbox_inches='tight')
'''


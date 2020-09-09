#################################
# load SCB_RIVERS.mat
# contains Rationale Method data
# turn certain rivers into netCDF4 data
# also turn certain Bight and USGS river data 
# into netCDF
#################################
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
import numpy as np
import scipy.io as sio
import datetime
import time
from netCDF4 import Dataset, num2date, date2num, stringtochar
import pickle

# dictionary with keys
# NO3, NH4, name, list (rational IDs), PO4, OP, ON, Flow
# data structured (time, river)
#                 (8785, 48)
rational_rivers = sio.loadmat('SCB_RIVERS.mat')

# grid path for lat/lon and grid calculations
#grid_path = '/data/project3/kesf/ROMS/L2_SCB/roms_grd.nc'

#################################################
# load final compilation data for nutrient data 
# from open_excel_rivers_no_watershed.py
################################################
river_comp_conv = pickle.load(open('river_data_conv_comp_no_watershed.pkl','rb'))
river_comp_conv_removed = pickle.load(open('river_data_conv_comp_removed.pkl','rb'))
time_data_c = np.load('river_time_data_comp.npy')

# get temperature data
temp_sm = np.load('temperature_santa_margarita.npy')
temp_la = np.load('temperature_los_angeles.npy')


# time is stored as days from 1/1/1990 to 1/1/2014
# (assuming 366 days, done by Faycal, end date will be later when converted to netCDF4)
time_data_r = range(rational_rivers['Flow'].shape[0])

# exclude index of these rivers because will not use rational method for data
exclude = [ 2,
            6,
            16,
            17,
            19,
            22,
            26,
            27,
            29,
            30,
            31,
            37,
            38,
            39,
            41,
            43,
            45,
    # exclude rivers with 0 flow as well and use Bight 08 data for it
            0,
            5,
            8,
            15,
            21,
            28,
            34,
            32,
            42
            ]

######################################
# get river names from rational rivers
######################################
river_names_r = []
for i,r in enumerate(rational_rivers['name'][0]):
    if i not in exclude:
        river_names_r.append(r[0])

river_names_r.append('Arroyo Burro')
river_names_r.append('Canada de la Gaviota')
river_names_r.append('Franklin Creek')

#######################################
# get river names from compilation data
########################################
river_names_c = [
                '154-San_Juan_Crk',
                '345-Goleta_SanJose',
                '350-Montecito',
                '237-SanDiegoR',
                '257-Sweetwater',
                '109-Solstice Canyon',
                '32-LARiver',
                '345-Goleta_Atascadero',
                '189-Salt Creek',
                '98-little Sycamore',
                '34-StaAnaRiver',
                '119-Pena Canyon',
                '177-Moro Canyon',
                '85-Ballona_Crk',
                '262-Tijuana',
                'Santa Margarita River',
                '143-LAHarbor',
                '116-Tuna Canyon',
                '351-Rincon',
                '317-Marie Canyon',
                '45-Santa_Clara',
                '91-Santa Monica Canyon',
                '210-Aliso Canyon',
                '108-Las Flores Canyon',
                '267-MissionBay',
                '288-Otay',
                '256-LPL',
                '331-Encinas',
                '95-Arroyo Sequit',
                '141-SanDiegoCrk',
                '7-VenturaRiv',
                '130-RedondoBchKingHarbor',
                '354-Mission Creek',
                '112-Walnut Canyon',
                '101-Trancas canyon',
                '111-Carbon Canyon',
                '37-Calleguas',
                '36-SanGabrielR',
                '201-SanLuisReyR',
                # missing flow values in rational river data
                '227-AguaHedionda', 
                '221-BuenaVista',
                '224-EscondidoCrk',
                '206-LasFlores',
                '217-SanDieguito',
                '199-SanOnofreCrk',
                '225-SanMarcosCrk',
                '279-TecoloteCrk',
                '287-Chollas-Crk'
                ]
pickle.dump(river_names_r,open('river_names_24.pkl','wb'))
pickle.dump(river_names_c,open('river_names_10.pkl','wb'))

##########################################
# get flow data for specific rivers needed 
# from Status of River Compilation 
##########################################
flow_data_r_include = np.zeros([len(time_data_r),len(rational_rivers['name'][0])+3])
# array excludes rivers that need data from final compilation (8785,21)
flow_data_r = np.zeros([len(time_data_r),len(rational_rivers['name'][0])-len(exclude)+3])
# get flow data 
for t in list(range(1,len(time_data_r))):
    for r in range(len(rational_rivers['name'][0])):
        if r not in exclude: 
            flow_data_r_include[t-1,r] = rational_rivers['Flow'][t,r+1]

# remove empty arrays and have the flow_data_r indexes match up with the river_names_r
for t_f in list(range(1,len(time_data_r))): 
    flow_data_r[t_f-1] = np.delete(flow_data_r_include[t_f-1],exclude)  

#tranpose flow_data_r to (river,time) (21,8785)
flow_data_r = np.transpose(flow_data_r,(1,0))

##############################
# get nutrient data 
# from final river compilation 
# and assign dry/wet weather values 
# shape (75,3653)
##############################
# reshape final compilation data to individual arrays per variable
# data structure
# 0 latitude
# 1 longitude
# 2 flow
# 3 total nitrogen
# 4 total phosphorous
# 5 ammonia
# 6 nitrate
# 7 phosphate
flow_data_c = np.zeros([len(river_names_c),len(time_data_c)])
TN_data_c = np.zeros([len(river_names_c),len(time_data_c)])
TP_data_c = np.zeros([len(river_names_c),len(time_data_c)])
NH4_data_c = np.zeros([len(river_names_c),len(time_data_c)])
NO3_data_c = np.zeros([len(river_names_c),len(time_data_c)])
PO4_data_c = np.zeros([len(river_names_c),len(time_data_c)])
lat_data_c = np.zeros(len(river_names_c))
lon_data_c = np.zeros(len(river_names_c))
alk_data_c = np.zeros([len(river_names_c),len(time_data_c)])
temp_data_c = np.zeros([len(river_names_c),len(time_data_c)])

# r_i and r_c ensures order is same as river_names_c
for r_i,r_c in enumerate(river_names_c):
    lat_data_c[r_i] = river_comp_conv[r_c][0]
    lon_data_c[r_i] = river_comp_conv[r_c][1]
    for t_c in range(len(time_data_c)):
        flow_data_c[r_i,t_c] = river_comp_conv[r_c][2][t_c]
        TN_data_c[r_i,t_c] = river_comp_conv[r_c][3][t_c]
        TP_data_c[r_i,t_c] = river_comp_conv[r_c][4][t_c]
        NH4_data_c[r_i,t_c] = river_comp_conv[r_c][5][t_c]
        NO3_data_c[r_i,t_c] = river_comp_conv[r_c][6][t_c]
        PO4_data_c[r_i,t_c] = river_comp_conv[r_c][7][t_c]
        alk_data_c[r_i,t_c] = river_comp_conv[r_c][8][t_c]
        temp_data_c[r_i,t_c] = river_comp_conv[r_c][9][t_c]


#############################
# nutrient data for rational
#############################
# change shape to have (rivers,time) (30,8785)
# add 3 for Arroyo Burro, Canada de la Gaviota, and Franklin Creek
TN_data_r = np.zeros([len(flow_data_r),len(time_data_r)])
TP_data_r = np.zeros([len(flow_data_r),len(time_data_r)])
NH4_data_r = np.zeros([len(flow_data_r),len(time_data_r)])
NO3_data_r = np.zeros([len(flow_data_r),len(time_data_r)])
PO4_data_r = np.zeros([len(flow_data_r),len(time_data_r)])
lat_data_r = np.zeros(len(flow_data_r))
lon_data_r = np.zeros(len(flow_data_r))
alk_data_r = np.zeros([len(flow_data_r),len(time_data_r)])
temp_data_r = np.zeros([len(flow_data_r),len(time_data_r)])
           
time_units_r = 'days since 1990-01-01'
time_data_r_date = num2date(time_data_r,units=time_units_r)

# get list of rational rivers to append nutrient and lat/lon data  
# same order as river_names_r
remove = [  #'227-AguaHedionda',
            '154-Arroyo_Trabuco',
            '34-BolsaChicaWestminster',
            '141-BonitaCrk',
            #'221-BuenaVista',
            '352-Carpinteria', 
            #'287-Chollas-Crk', 
            '141-CostaMesaChnl',
            '36-CoyoteCrk',
            '174-Cristianitos_Crk',
            '345-DevLagoon',
            '131-Dominguez',
            '34-EGardenGroveWinter', 
            #'224-EscondidoCrk',
            '345-Goleta_Tecolotito',
            '176-LagunaCyn',
            #'206-LasFlores',
            '71-MalibuCrk',
            '192-Prima_Desch',
            '37-Revolon',
            #'217-SanDieguito',
            #'225-SanMarcosCrk',
            '174-San Mateo',
            #'199-SanOnofreCrk',
            '345-San_Pedro_Crk',
            '141-StaAnaDelhi',
            '193-Segunda_Desch',
            #'279-TecoloteCrk',
            '86-Topanga',
            '103-Zuma Canyon',
            'Arroyo Burro Creek',
            'Canada de la Gaviota',
            'Franklin Creek' ]

# range of dry dates for each year (May 1 to Oct 31)
dry = range(121,305) 

# append nutrient data to each variable and index of correct river
for r in range(len(river_comp_conv_removed.keys())-1):
    lat_data_r[r] = river_comp_conv_removed[remove[r]][0]
    lon_data_r[r] = river_comp_conv_removed[remove[r]][1]
    for t_n,d in enumerate(time_data_r_date):
        alk_data_r[r,t_n] = river_comp_conv_removed[remove[r]][8][0]
        # t_n%365 to repeat over the same year for temperature
        temp_data_r[r,t_n] = river_comp_conv_removed[remove[r]][9][t_n%365]
        if datetime.datetime.timetuple(d).tm_yday in dry:
            TN_data_r[r,t_n] = river_comp_conv_removed[remove[r]][3][150]
            TP_data_r[r,t_n] = river_comp_conv_removed[remove[r]][4][150]
            NH4_data_r[r,t_n] = river_comp_conv_removed[remove[r]][5][150]
            NO3_data_r[r,t_n] = river_comp_conv_removed[remove[r]][6][150]
            PO4_data_r[r,t_n] = river_comp_conv_removed[remove[r]][7][150]
        else:
            TN_data_r[r,t_n] = river_comp_conv_removed[remove[r]][3][0]
            TP_data_r[r,t_n] = river_comp_conv_removed[remove[r]][4][0]
            NH4_data_r[r,t_n] = river_comp_conv_removed[remove[r]][5][0]
            NO3_data_r[r,t_n] = river_comp_conv_removed[remove[r]][6][0]
            PO4_data_r[r,t_n] = river_comp_conv_removed[remove[r]][7][0]

###########################################
# SPECIAL CASE FOR Arroyo Burro,
# Canada de la Gaviota, and Franklin Creek
# load and convert data
##########################################
# data structure 
# 0 latitude
# 1 longitude
# 2 flow
# 3 total nitrogen
# 4 total phosphorous
# 5 ammonia
# 6 nitrate
# 7 phosphate
# 8 alkalinity
# 9 temperature

# convert g/day,mg/L to mmol/m3
# convert m3/day to m3/s
day_to_second = 1./86400
cc_mg_to_mmol = 1./100.09
N_mg_to_mmol = 1./14
P_mg_to_mmol = 1./30.97
L_to_m3 = 1000
g_to_mg = 1000

wb2 = openpyxl.load_workbook('extra_rivers.xlsx',data_only=True)
extra_rivers_names = wb2.sheetnames

extra_rivers_data = defaultdict(list)
for r_e in extra_rivers_names:
    extra_rivers_data[r_e] = [ [] for i in range(9)]

for r_e,name_e in enumerate(extra_rivers_names): 
    lat_data_r[r_e-3] = wb2[name_e].cell(row=2,column=6).value
    lon_data_r[r_e-3] = wb2[name_e].cell(row=2,column=7).value 
    temp_data_r[r_e-3] = temp_data_r[0]
    for row_e in range(2,8786):
        print('appending extra river data for row '+str(row_e)+' in '+str(8786))
        flow_data_r[r_e-3,row_e-2] = wb2[name_e].cell(row=row_e,column=2).value*day_to_second
        alk_data_r[r_e-3,row_e-2] = wb2[name_e].cell(row=2,column=10).value*cc_mg_to_mmol*L_to_m3
        if wb2[name_e].cell(row=row_e,column=2).value == 0:
            TN_data_r[r_e-3,row_e-2] = 0
            TP_data_r[r_e-3,row_e-2] = 0
            NH4_data_r[r_e-3,row_e-2] = 0
            NO3_data_r[r_e-3,row_e-2] = 0
            PO4_data_r[r_e-3,row_e-2] = 0
        if wb2[name_e].cell(row=row_e,column=2).value != 0:
            TN_data_r[r_e-3,row_e-2] = wb2[name_e].cell(row=2,column=8).value*N_mg_to_mmol*L_to_m3
            TP_data_r[r_e-3,row_e-2] = wb2[name_e].cell(row=2,column=9).value*P_mg_to_mmol*L_to_m3

            NH4_data_r[r_e-3,row_e-2] = wb2[name_e].cell(row=row_e,column=3).value*(1./wb2[name_e].cell(row=row_e,column=2).value)*g_to_mg*N_mg_to_mmol
            NO3_data_r[r_e-3,row_e-2] = wb2[name_e].cell(row=row_e,column=4).value*(1./wb2[name_e].cell(row=row_e,column=2).value)*g_to_mg*N_mg_to_mmol
            PO4_data_r[r_e-3,row_e-2] = wb2[name_e].cell(row=row_e,column=5).value*(1./wb2[name_e].cell(row=row_e,column=2).value)*g_to_mg*P_mg_to_mmol

'''
#########################################################
# MOVE RIVER LOCATIONS ONTO FIRST CELL IN OCEAN OFF LAND
# find grid cell values of the lat/lon, index into mask 
# to see if it is on land or water
# change lat/lon values to be first cell in ocean off land
#########################################################
# load grid
grid_nc = Dataset(grid_path,'r')
mask_nc = grid_nc.variables['mask_rho']
mask = np.copy(mask_nc)
#mask[mask==0.0] = np.nan
lat_nc = grid_nc.variables['lat_rho']
lon_nc = grid_nc.variables['lon_rho']
Ly = mask.shape[0] 
Lx = mask.shape[1] 
pm = grid_nc.variables['pm'][:,:]
pn = grid_nc.variables['pn'][:,:]
h = grid_nc.variables['h'][:,:]
'''
'''
# convert lat/lon coordinates to grid coordinates
river_coord_x_10 = []
river_coord_y_10 = []
river_coord_x_24 = []
river_coord_y_24 = []

for coord in range(len(lat_data_c)):
    lat_you_want = lat_data_c[coord]
    lon_you_want = lon_data_c[coord]
    temp = np.abs( (lat_nc - lat_you_want)**2 + (lon_nc - lon_you_want)**2)
    eta_coord,xi_coord = np.unravel_index(temp.argmin(),temp.shape)
    river_coord_x_10.append(xi_coord)
    river_coord_y_10.append(eta_coord)

for coord in range(len(lat_data_r)):
    lat_you_want = lat_data_r[coord]
    lon_you_want = lon_data_r[coord]
    temp = np.abs( (lat_nc - lat_you_want)**2 + (lon_nc - lon_you_want)**2)
    eta_coord,xi_coord = np.unravel_index(temp.argmin(),temp.shape)
    river_coord_x_24.append(xi_coord)
    river_coord_y_24.append(eta_coord)

# move river locations that are in ocean to 
# first grid cell on land 
# (next for loop ensures it is on first cell off land in ocean)
for i in range(len(river_coord_x_10)):
    while mask[river_coord_y_10[i],river_coord_x_10[i]] == 1:
        river_coord_x_10[i] += 1     

for i in range(len(river_coord_x_24)):
    while mask[river_coord_y_24[i],river_coord_x_24[i]] == 1:
        river_coord_x_24[i] += 1     
     
# move river locations that are on land offshore in x direction
# to first grid cell in ocean
for i in range(len(river_coord_x_10)):
    while mask[river_coord_y_10[i],river_coord_x_10[i]] == 0:
        river_coord_x_10[i] -= 1 

for i in range(len(river_coord_x_24)):
    while mask[river_coord_y_24[i],river_coord_x_24[i]] == 0:
        river_coord_x_24[i] -= 1 

# check if any river locations are still on land 
n = 0
for i in range(len(river_coord_x_24)):
    if mask[river_coord_y_24[i],river_coord_x_24[i]] == 0: 
        n+=1

for i in range(len(river_coord_x_10)):
    if mask[river_coord_y_10[i],river_coord_x_10[i]] == 0:
        n+=1

# convert correct x and y grid cells to lat/lon for netcdf
for c in range(len(river_coord_x_10)):
    lat_data_c[c] = lat_nc[river_coord_y_10[c],river_coord_x_10[c]]
    lon_data_c[c] = lon_nc[river_coord_y_10[c],river_coord_x_10[c]]

for c in range(len(river_coord_x_24)):
    lat_data_r[c] = lat_nc[river_coord_y_24[c],river_coord_x_24[c]]
    lon_data_r[c] = lon_nc[river_coord_y_24[c],river_coord_x_24[c]]

'''
'''
###########################################
# CREATE NETCDF FILES FOR SOUTH COAST RIVERS
###########################################              
# create netCDF files, one for bight/usgs data and one for rational method data
f0 = Dataset('south_coast_rivers_10_years_no_watershed_new.nc','w')
f1 = Dataset('south_coast_rivers_24_years_new.nc','w')

# details about bight/usgs data
f0.title = 'Southern California Bight Coastal River Data, 10 years of data'
f0.source = 'Southern California Bight \'08 Data, USGS Gauge data'
f0.history = 'Created '+time.ctime(time.time())
f0.description = 'Rivers in this data set: 154-San_Juan_Crk, 345-Goleta_SanJose, 350-Montecito, 237-SanDiegoR, 257-Sweetwater, 109-Solstice Canyon, 32-LARiver, 345-Goleta_Atascadero, 189-Salt Creek, 98-little Sycamore, 34-StaAnaRiver, 119-Pena Canyon, 177-Moro Canyon, 85-Ballona_Crk, 262-Tijuana, Santa Margarita River, 143-LAHarbor, 116-Tuna Canyon, 351-Rincon, 317-Marie Canyon, 45-Santa_Clara, 91-Santa Monica Canyon, 210-Aliso Canyon, 108-Las Flores Canyon, 267-MissionBay, 288-Otay, 256-LPL(Los Penasquitos), 331-Encinas, 95- Arroyo Sequit, 141-SanDiegoCrk, 7-VenturaRiv, 130-RedondoBchKingHarbor, 354-Mission Creek, 112-Walnut Canyon, 101-Trancas canyon, 111-Carbon Canyon, 37-Calleguas, 36-SanGabrielR, 201-SanLuisReyR, 227-AguaHedionda, 221-BuenaVista, 224-EscondidoCrk, 206-LasFlores, 217-SanDieguito, 199-SanOnofreCrk, 225-SanMarcosCrk, 279-TecoloteCrk, 287-Chollas-Crk' 


f1.title = 'Southern California Bight Coastal River Data, 24 years of data'
f1.source = 'Rational Methods, Ashmita Sengupta\'s model'
f1.history = 'Created '+time.ctime(time.time())
f1.description = 'Rivers in this data set: Arroyo Trabuco, Bolsa Chica Westminster Channel, Bonita Creek, Carpinteria, Costa Mesa Chanel, Coyote Creek, Cristianitos Creek, Devereux Lagoon, Dominguez, E Garden Grove Wintersberg Channel, Goleta Tecolotito, Laguna Canyon, Malibu Creek, Prima Desch, Revolon, San Mateo, San Pedro Creek, Santa Ana Delhi, Segunda Desch, Topanga, Zuma Canyon, Arroyo Burro, Canada de la Gaviota, Franklin Creek'


# create time, lat, lon dimensions
time_c = f0.createDimension('time',None)
lat_c = f0.createDimension('lat',len(river_names_c))
lon_c = f0.createDimension('lon',len(river_names_c))

time_r = f1.createDimension('time',None)
lat_r = f1.createDimension('lat',len(river_names_r))
lon_r = f1.createDimension('lon',len(river_names_r))


# create variables
times_c = f0.createVariable('time',np.float64,('time',))
lats_c = f0.createVariable('latitude',np.float32,('lat',))
lons_c = f0.createVariable('longitude',np.float32,('lon',))

flow_c = f0.createVariable('flow',np.float32,('time','lat','lon'))
NH4_c = f0.createVariable('ammonium',np.float32,('time','lat','lon'))
NO3_c = f0.createVariable('nitrate',np.float32,('time','lat','lon'))
PO4_c = f0.createVariable('phosphate',np.float32,('time','lat','lon'))
TN_c = f0.createVariable('total_nitrogen',np.float32,('time','lat','lon'))
TP_c = f0.createVariable('total_phosphorus',np.float32,('time','lat','lon'))
alk_c = f0.createVariable('alkalinity',np.float32,('time','lat','lon'))
temp_c = f0.createVariable('temperature',np.float32,('time','lat','lon'))


times_r = f1.createVariable('time',np.float64,('time',))
lats_r = f1.createVariable('latitude',np.float32,('lat',))
lons_r = f1.createVariable('longitude',np.float32,('lon',))

flow_r = f1.createVariable('flow',np.float32,('time','lat','lon'))
NH4_r = f1.createVariable('ammonium',np.float32,('time','lat','lon'))
NO3_r = f1.createVariable('nitrate',np.float32,('time','lat','lon'))
PO4_r = f1.createVariable('phosphate',np.float32,('time','lat','lon'))
TN_r = f1.createVariable('total_nitrogen',np.float32,('time','lat','lon'))
TP_r = f1.createVariable('total_phosphorus',np.float32,('time','lat','lon'))
alk_r = f1.createVariable('alkalinity',np.float32,('time','lat','lon'))
temp_r = f1.createVariable('temperature',np.float32,('time','lat','lon'))

# units
times_c.units = 'days since 1997-1-1'
lats_c.units = 'degrees north'
lons_c.units = 'degrees east'
flow_c.units = 'm3/s'
NH4_c.units = 'mmol/m3'
PO4_c.units = 'mmol/m3'
NO3_c.units = 'mmol/m3'
TN_c.units = 'mmol/m3'
TP_c.units = 'mmol/m3'
alk_c.units = 'mmol/m3'
temp_c.units = 'degrees Celsius'

times_r.units = 'days since 1990-1-1'
lats_r.units = 'degrees north'
lons_r.units = 'degrees east'
flow_r.units = 'm3/s'
NH4_r.units = 'mmol/m3'
PO4_r.units = 'mmol/m3'
NO3_r.units = 'mmol/m3'
TN_r.units = 'mmol/m3'
TP_r.units = 'mmol/m3'
alk_r.units = 'mmol/m3'
temp_r.units = 'degrees Celsius'

# assign data
times_c[:] = date2num(time_data_c,times_c.units)
lats_c[:] = lat_data_c
lons_c[:] = lon_data_c

for t_c in range(len(time_data_c)):
    for lat in range(len(lat_data_c)):
        flow_c[t_c,lat,lat] = flow_data_c[lat,t_c]
        NH4_c[t_c,lat,lat] = NH4_data_c[lat,t_c]
        NO3_c[t_c,lat,lat] = NO3_data_c[lat,t_c]
        PO4_c[t_c,lat,lat] = PO4_data_c[lat,t_c]
        TN_c[t_c,lat,lat] = TN_data_c[lat,t_c]
        TP_c[t_c,lat,lat] = TP_data_c[lat,t_c]
        alk_c[t_c,lat,lat] = alk_data_c[lat,t_c]
        temp_c[t_c,lat,lat] = temp_data_c[lat,t_c]
    print('assigning data from Compilation time '+str(t_c)+' of '+str(len(time_data_c)))
        
times_r[:] = np.asarray(time_data_r)
lats_r[:] = lat_data_r
lons_r[:] = lon_data_r

for t_r in range(len(time_data_r)):
    for lon in range(len(lon_data_r)):
        flow_r[t_r,lon,lon] = flow_data_r[lon,t_r]
        NH4_r[t_r,lon,lon] = NH4_data_r[lon,t_r]
        NO3_r[t_r,lon,lon] = NO3_data_r[lon,t_r]
        PO4_r[t_r,lon,lon] = PO4_data_r[lon,t_r]
        TN_r[t_r,lon,lon] = TN_data_r[lon,t_r]
        TP_r[t_r,lon,lon] = TP_data_r[lon,t_r]
        alk_r[t_r,lon,lon] = alk_data_r[lon,t_r]
        temp_r[t_r,lon,lon] = temp_data_r[lon,t_r]
    print('assigning data from Rational time '+str(t_r)+' of '+str(len(time_data_r)))

f0.close()
f1.close()
print('Compilation data written to south_coast_rivers_10_years_no_watershed_new.nc')
print('Rational data written to south_coast_rivers_24_years_new.nc')
'''

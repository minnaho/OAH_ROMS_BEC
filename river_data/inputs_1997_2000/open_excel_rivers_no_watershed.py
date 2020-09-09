##################################################
# take in excel data from
# rivers 
# and put into .nc file
# structure: 
##################################################
import openpyxl
from openpyxl import Workbook
from netCDF4 import Dataset
import datetime
import time
import numpy as np
from collections import defaultdict
from netCDF4 import num2date, date2num
import pickle

# load excel file
wb1 = openpyxl.load_workbook('Final River Compilation latlon updated nowatersheds.xlsx',data_only=True)

# each sheet is a river in the excel file
river = wb1.sheetnames

# make lists for each input to append data into
time_data = []

# append time
# 3653 = amount of days from 1/1/1997 to 1/1/2007
# same for every river
for t in range(2,3652): 
    time_data.append(wb1[river[0]].cell(row=t,column=1).value)   

# make dictionary of river data
river_data = defaultdict(list)
for r in river:
    river_data[r] = [ [] for i in range(10)]
        

# data structure 
# 0 latitude
# 1 longitude
# 2 flow
# 3 total nitrogen
# 4 total phosphorous
# 5 ammonia
# 6 nitrate
# 7 phosphate
# 8 alkalinity

# append lat/lon, and alkalinity
for r in river:
    river_data[r][0].append(wb1[r].cell(row=2,column=28).value)
    river_data[r][1].append(wb1[r].cell(row=2,column=29).value)
    river_data[r][8].append(wb1[r].cell(row=2,column=14).value)


# tuple days from May 1 - Oct 31
dry = range(121,305)

# append flow and nutrient data 
for r in river:  
    for i,d in enumerate(time_data):
        # check if date is in not dry period or if there is an empty cell and append appropriate cell value
        if datetime.datetime.timetuple(d).tm_yday not in dry or wb1[r].cell(row=2,column=23).value == None:
            river_data[r][3].append(wb1[r].cell(row=2,column=18).value)
            river_data[r][4].append(wb1[r].cell(row=2,column=19).value)
            river_data[r][5].append(wb1[r].cell(row=2,column=20).value)
            river_data[r][6].append(wb1[r].cell(row=2,column=21).value)
            river_data[r][7].append(wb1[r].cell(row=2,column=22).value)
        else:
            river_data[r][3].append(wb1[r].cell(row=2,column=23).value)
            river_data[r][4].append(wb1[r].cell(row=2,column=24).value)
            river_data[r][5].append(wb1[r].cell(row=2,column=25).value)
            river_data[r][6].append(wb1[r].cell(row=2,column=26).value)
            river_data[r][7].append(wb1[r].cell(row=2,column=27).value)
        # take in flow data if there is a value in that cell
        if wb1[r].cell(row=i+2,column=2).value == None or wb1[r].cell(row=i+2,column=2).value=='':
            if wb1[r].cell(row=i+2,column=3).value == None or wb1[r].cell(row=i+2,column=3).value=='':
                river_data[r][2].append(wb1[r].cell(row=i+2,column=4).value)    
            elif wb1[r].cell(row=i+2,column=4).value == None or wb1[r].cell(row=i+2,column=4).value=='':
                river_data[r][2].append(wb1[r].cell(row=i+2,column=3).value)    

        elif wb1[r].cell(row=i+2,column=2).value != None or wb1[r].cell(row=i+2,column=2).value != '':
            river_data[r][2].append(wb1[r].cell(row=i+2,column=2).value)

        # add constant alkalinity value for every river
        river_data[r][8].append(wb1[r].cell(row=2,column=14).value) 

        print('appending data for date '+str(d)+' of '+str(time_data[-1]))

# MAKE RIVER FLOW DATA ONLY 1 YEAR OF FLOWS
# exclude rivers that use USGS data and unique data
usgs_rivers = ['Santa Margarita River','237-SanDiegoR','37-Calleguas','45-Santa_Clara','256-LPL','345-Goleta_SanJose','7-VenturaRiv','154-San_Juan_Crk','345-Goleta_Atascadero']
for r in river:
    if r not in usgs_rivers:
        for i in range(len(time_data)):
            river_data[r][2] = river_data[r][2][:365]*10

# special cases where nutrients are stored differently for these 2 rivers
for n in range(3,8):
    river_data['32-LARiver'][n] = [] 
    river_data['36-SanGabrielR'][n] = [] 
 
for row0 in range(2,3655):      
    river_data['32-LARiver'][3].append(wb1['32-LARiver'].cell(row=row0,column=18).value)
    river_data['32-LARiver'][4].append(wb1['32-LARiver'].cell(row=row0,column=11).value) 
    river_data['32-LARiver'][5].append(wb1['32-LARiver'].cell(row=row0,column=7).value)
    river_data['32-LARiver'][6].append(wb1['32-LARiver'].cell(row=row0,column=5).value)
    river_data['32-LARiver'][7].append(wb1['32-LARiver'].cell(row=row0,column=10).value) 
    river_data['36-SanGabrielR'][3].append(wb1['36-SanGabrielR'].cell(row=row0,column=18).value)
    river_data['36-SanGabrielR'][4].append(wb1['36-SanGabrielR'].cell(row=row0,column=11).value) 
    river_data['36-SanGabrielR'][5].append(wb1['36-SanGabrielR'].cell(row=row0,column=7).value)
    river_data['36-SanGabrielR'][6].append(wb1['36-SanGabrielR'].cell(row=row0,column=5).value)
    river_data['36-SanGabrielR'][7].append(wb1['36-SanGabrielR'].cell(row=row0,column=10).value)


#####################################
# convert data from ft3/s to m3/s
# convert data from mg/L to mmol/m3
####################################
ft3_to_m3 = 0.02831685
L_to_m3 = 1./.001
N_mg_to_mmol = 1./14
P_mg_to_mmol = 1./30.97
# calcium carbonate mg to mmol
cc_mg_to_mmol = 1./100.09


# make dictionary of converted river data
river_data_conv = defaultdict(list)
for r in river:
    river_data_conv[r] = [ [] for i in range(10)]

###################
# add temperature 
###################
# concrete channels use temperature from LA river
temp_la = np.load('temperature_los_angeles.npy')
temp_la_names = ['237-SanDiegoR','32-LARiver','34-StaAnaRiver','85-Ballona_Crk','141-SanDiegoCrk','7-VenturaRiv','37-Calleguas','36-SanGabrielR','201-SanLuisReyR','154-San_Juan_Crk','34-BolsaChicaWestminster']

# nonconcrete rivers use temperature from Santa margarita river
temp_sm = np.load('temperature_santa_margarita.npy')
 
for r in river:
    # change None values to nan by changing to float array 
    # (also avoids a TypeError of multiplying None by float)
    river_data_conv[r][0] = river_data[r][0][0]
    river_data_conv[r][1] = river_data[r][1][0]
    river_data_conv[r][2] = np.asarray(river_data[r][2],float) * ft3_to_m3
    river_data_conv[r][3] = np.asarray(river_data[r][3],float) * L_to_m3 * N_mg_to_mmol
    river_data_conv[r][4] = np.asarray(river_data[r][4],float) * L_to_m3 * P_mg_to_mmol
    river_data_conv[r][5] = np.asarray(river_data[r][5],float) * L_to_m3 * N_mg_to_mmol
    river_data_conv[r][6] = np.asarray(river_data[r][6],float) * L_to_m3 * N_mg_to_mmol
    river_data_conv[r][7] = np.asarray(river_data[r][7],float) * L_to_m3 * P_mg_to_mmol
    river_data_conv[r][8] = np.asarray(river_data[r][8],float) * L_to_m3 * cc_mg_to_mmol
    if r in temp_la_names: 
        river_data_conv[r][9] = np.array(list(temp_la[:365])*10)
    else:
        river_data_conv[r][9] = np.array(list(temp_sm[:365])*10)

###########################################################
# remove rivers from this data set used in rational method 
###########################################################
remove = [  #'227-AguaHedionda',
            '103-Zuma Canyon',
            #'221-BuenaVista',
            '141-StaAnaDelhi',
            '345-San_Pedro_Crk',
            '174-Cristianitos_Crk',
            '34-EGardenGroveWinter',
            '345-Goleta_Tecolotito',
            '71-MalibuCrk',
            #'224-EscondidoCrk',
            #'206-LasFlores',
            #'217-SanDieguito',
            '141-CostaMesaChnl',
            '141-BonitaCrk',
            '345-DevLagoon',
            #'199-SanOnofreCrk',
            '154-Arroyo_Trabuco',
            '174-San Mateo',
            #'225-SanMarcosCrk',
            '86-Topanga',
            #'279-TecoloteCrk',
            '37-Revolon',
            '193-Segunda_Desch',
            '352-Carpinteria',
            '36-CoyoteCrk',
            '34-BolsaChicaWestminster',
            '192-Prima_Desch',
            '131-Dominguez',
            #'287-Chollas-Crk',  
            '176-LagunaCyn',
            '153-Santa Margarita'     ]
    
        
river_data_conv_removed = defaultdict(list) 
for i in remove:
    river_data_conv_removed[i] = river_data_conv[i]
    river_data_conv.pop(i)
    print('removed '+str(i)+' in river_data_conv')
 
# save data into .npy files to call later and save processing time
np.save('river_time_data_comp.npy',time_data)
pickle.dump(river_data,open('river_data_comp_no_watershed.pkl','wb'))
pickle.dump(river_data_conv,open('river_data_conv_comp_no_watershed.pkl','wb'))
pickle.dump(river_data_conv_removed,open('river_data_conv_comp_removed.pkl','wb'))


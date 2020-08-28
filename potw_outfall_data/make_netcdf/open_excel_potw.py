######################################################
# TURN EXCEL DATA INTO DICTIONARY OF DATA IN PYTHON
# POTW_INTERP-ms edits_Minna_edits.xlsx data into a dictionary
# 0 dimension for each key is date
# 1 dimension is flowrate m^3/s 
# 2 dimension is NO3 in mg/L 
# 3 dimension is NO2 in mg/L 
# 4 dimension is NH4 in mg/L 
# 5              BOD    mg/L 
# 6              COD    mg/L 
# 7              Fe     mg/L 
# 8              SO2    mg/L 
# 9              ON     mg/L (organic nitrogen?)
# 10             TP     mg/L  
# 11             PO4    mg/L
# 12             OP     mg/L
# 13             TOC    mg/L
######################################################

import openpyxl
from openpyxl import Workbook
from collections import defaultdict
import pickle
import numpy as np

wb = openpyxl.load_workbook('POTW_INTERP-ms edits_Minna_edits_converted.xlsx',data_only=True)
location = wb.sheetnames

potw_data = defaultdict(list)

for i in range(len(location)):
    potw_data[location[i]] = [ [] for i in range(14)]

for j in range(len(location)):
    for i in range(2,559):
        date_i = wb[location[j]].cell(row=i,column=1).value
        flow_i = wb[location[j]].cell(row=i,column=2).value
        NO3_i = wb[location[j]].cell(row=i,column=3).value
        NO2_i = wb[location[j]].cell(row=i,column=4).value
        NH4_i = wb[location[j]].cell(row=i,column=5).value
        BOD_i = wb[location[j]].cell(row=i,column=6).value
        COD_i = wb[location[j]].cell(row=i,column=7).value
        Fe_i = wb[location[j]].cell(row=i,column=8).value
        SO2_i = wb[location[j]].cell(row=i,column=9).value
        ON_i = wb[location[j]].cell(row=i,column=10).value
        TP_i = wb[location[j]].cell(row=i,column=11).value
        PO4_i = wb[location[j]].cell(row=i,column=12).value
        OP_i = wb[location[j]].cell(row=i,column=13).value
        TOC_i = wb[location[j]].cell(row=i,column=14).value
     
        potw_data[location[j]][0].append(date_i)
        potw_data[location[j]][1].append(flow_i)
        potw_data[location[j]][2].append(NO3_i)
        potw_data[location[j]][3].append(NO2_i)
        potw_data[location[j]][4].append(NH4_i)
        potw_data[location[j]][5].append(BOD_i) 
        potw_data[location[j]][6].append(COD_i) 
        potw_data[location[j]][7].append(Fe_i)
        potw_data[location[j]][8].append(SO2_i)
        potw_data[location[j]][9].append(ON_i)
        potw_data[location[j]][10].append(TP_i)
        potw_data[location[j]][11].append(PO4_i)
        potw_data[location[j]][12].append(OP_i)
        potw_data[location[j]][13].append(TOC_i)
                                             
 
    print('Adding data to dict for location ' + str(location[j]))

pickle.dump(potw_data,open('potw_data.pkl','wb'))

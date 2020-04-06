################################
# make netCDF4 file of POTW data
# use ncdump -h potw_data_converted.nc
# to see details of file
###############################
from netCDF4 import Dataset
import pickle
import numpy as np
import datetime
from netCDF4 import num2date, date2num
import openpyxl

# load data from dictionary that was created 
# in plot_potw.py
potw_data_converted = pickle.load(open('potw_data_converted.pkl','rb'))
site_sum = openpyxl.load_workbook('../POTW Outfall Site Summary.xlsx')

# grid path
#grid_path = '../../usw1_grd.nc' 

outfalls = ['Hyperion','JWPCP','OCSD','PLWTP']

###############
# LOAD GRID
###############
# load grid
'''
grid_nc = Dataset(grid_path,'r')
mask_nc = grid_nc.variables['mask_rho']
mask = np.copy(mask_nc)
#mask[mask==0.0] = np.nan
lat_nc = grid_nc.variables['lat_rho']
lon_nc = grid_nc.variables['lon_rho']
Ly = mask.shape[0] # shape 1410
Lx = mask.shape[1] # shape 770
pm = grid_nc.variables['pm'][:,:]
pn = grid_nc.variables['pn'][:,:]
h = grid_nc.variables['h'][:,:]
'''
################################
# get lat/lon of major outfalls
# and pH and sulfate and alkalinity
# from site summary
################################
hyp_lat = site_sum['Sheet1'].cell(row=14,column=4).value
hyp_lon = site_sum['Sheet1'].cell(row=14,column=5).value
hyp_alk = np.full((len(potw_data_converted['Hyperion'][1])),site_sum['Sheet1'].cell(row=14,column=13).value)
hyp_pH = np.full((len(potw_data_converted['Hyperion'][1])),site_sum['Sheet1'].cell(row=14,column=14).value)
hyp_SiO4 = np.full((len(potw_data_converted['Hyperion'][1])),site_sum['Sheet1'].cell(row=14,column=15).value)
 
hyp_temp = np.full((len(potw_data_converted['Hyperion'][1])),site_sum['Sheet1'].cell(row=14,column=16).value)
hyp_DO = np.full((len(potw_data_converted['Hyperion'][1])),site_sum['Sheet1'].cell(row=14,column=17).value)
hyp_salinity = np.full((len(potw_data_converted['Hyperion'][1])),site_sum['Sheet1'].cell(row=14,column=18).value)

jwpcp_lat = site_sum['Sheet1'].cell(row=12,column=4).value
jwpcp_lon = site_sum['Sheet1'].cell(row=12,column=5).value
jwpcp_alk = np.full((len(potw_data_converted['JWPCP'][1])),site_sum['Sheet1'].cell(row=12,column=13).value)
jwpcp_pH = np.full((len(potw_data_converted['JWPCP'][1])),site_sum['Sheet1'].cell(row=12,column=14).value)
jwpcp_SiO4 = np.full((len(potw_data_converted['JWPCP'][1])),site_sum['Sheet1'].cell(row=12,column=15).value)
jwpcp_temp = np.full((len(potw_data_converted['JWPCP'][1])),site_sum['Sheet1'].cell(row=12,column=16).value)
jwpcp_DO = np.full((len(potw_data_converted['JWPCP'][1])),site_sum['Sheet1'].cell(row=12,column=17).value)
jwpcp_salinity = np.full((len(potw_data_converted['JWPCP'][1])),site_sum['Sheet1'].cell(row=12,column=18).value)

ocsd_lat = site_sum['Sheet1'].cell(row=11,column=4).value
ocsd_lon = site_sum['Sheet1'].cell(row=11,column=5).value
ocsd_alk = np.full((len(potw_data_converted['OCSD'][1])),site_sum['Sheet1'].cell(row=11,column=13).value)
ocsd_pH = np.full((len(potw_data_converted['OCSD'][1])),site_sum['Sheet1'].cell(row=11,column=14).value)
ocsd_SiO4 = np.full((len(potw_data_converted['OCSD'][1])),site_sum['Sheet1'].cell(row=11,column=15).value)
ocsd_temp = np.full((len(potw_data_converted['OCSD'][1])),site_sum['Sheet1'].cell(row=11,column=16).value)
ocsd_DO = np.full((len(potw_data_converted['OCSD'][1])),site_sum['Sheet1'].cell(row=11,column=17).value)
ocsd_salinity = np.full((len(potw_data_converted['OCSD'][1])),site_sum['Sheet1'].cell(row=11,column=18).value)

plwtp_lat = site_sum['Sheet1'].cell(row=2,column=4).value
plwtp_lon = site_sum['Sheet1'].cell(row=2,column=5).value
plwtp_alk = np.full((len(potw_data_converted['PLWTP'][1])),site_sum['Sheet1'].cell(row=2,column=13).value)
plwtp_pH = np.full((len(potw_data_converted['PLWTP'][1])),site_sum['Sheet1'].cell(row=2,column=14).value)
plwtp_SiO4 = np.full((len(potw_data_converted['PLWTP'][1])),site_sum['Sheet1'].cell(row=2,column=15).value)
plwtp_temp = np.full((len(potw_data_converted['PLWTP'][1])),site_sum['Sheet1'].cell(row=2,column=16).value)
plwtp_DO = np.full((len(potw_data_converted['PLWTP'][1])),site_sum['Sheet1'].cell(row=2,column=17).value)
plwtp_salinity = np.full((len(potw_data_converted['PLWTP'][1])),site_sum['Sheet1'].cell(row=2,column=18).value)

alk_data = [hyp_alk,jwpcp_alk,ocsd_alk,plwtp_alk]
SiO4_data = [hyp_SiO4,jwpcp_SiO4,ocsd_SiO4,plwtp_SiO4]
pH_data = [hyp_pH,jwpcp_pH,ocsd_pH,plwtp_pH]
temp_data = [hyp_temp,jwpcp_temp,ocsd_temp,plwtp_temp]
DO_data = [hyp_DO,jwpcp_DO,ocsd_DO,plwtp_DO]
salinity_data = [hyp_salinity,jwpcp_salinity,ocsd_salinity,plwtp_salinity]

# convert alkalinity and SiO4 to mmol/m3
L_to_m3 = 1000
mol_wt_CaCO3 = 100.09
mol_wt_S = 32.065
mol_wt_O = 16

alk_data_conv = [ [] for i in range(len(alk_data)) ]
SiO4_data_conv = [ [] for i in range(len(SiO4_data)) ]
DO_data_conv = [ [] for i in range(len(DO_data)) ]

for i in range(len(alk_data)):
    alk_data_conv[i] = alk_data[i]*L_to_m3*(1./mol_wt_CaCO3)
    SiO4_data_conv[i] = SiO4_data[i]*L_to_m3*(1./mol_wt_S)
    DO_data_conv[i] = DO_data[i]*L_to_m3*(1./mol_wt_O)

#############################################################
# find grid cell values of the lat/lon, index into mask
# to see if it is on land or water
#############################################################
lat_data = np.array([hyp_lat,jwpcp_lat,ocsd_lat,plwtp_lat])
lon_data = np.array([hyp_lon,jwpcp_lon,ocsd_lon,plwtp_lon])


'''
potw_coord_x = []
potw_coord_y = []

for coord in range(len(lat_data)):
    lat_you_want = lat_data[coord]
    lon_you_want = lon_data[coord]
    temp = np.abs( (lat_nc - lat_you_want)**2 + (lon_nc - lon_you_want)**2)
    eta_coord,xi_coord = np.unravel_index(temp.argmin(),temp.shape)
    potw_coord_x.append(xi_coord)
    potw_coord_y.append(eta_coord)

# check if any potw locations are on land
n = 0
for i in range(len(potw_coord_x)):
    if mask[potw_coord_y[i],potw_coord_x[i]] == 0:
        n+=1

potw_coords = np.array((potw_coord_y,potw_coord_x))
potw_lat_lon = np.array((lat_data,lon_data))
#np.save('major_potw_lat_lon.npy',potw_lat_lon)
'''
#########################
# create netCDF file
#########################
f = Dataset('major_potw_data.nc','w')

# details about data
f.title = 'Major Southern California Bight POTW (Hyperion, JWPCP, OCSD, PLWTP) interpolated data 1970-2017'
f.source = 'Dr. Martha Sutula from Southern California Coastal Water Research Project'
#f.comment = 'Outfall dimension of 4 indicates each outfall: 0 is Hyperion, 1 is JWPCP, 2 is OCSD, 3 is PLWTP'

# create time and outfall dimensions
# time variable is unlimited so can put 
# as much data in there as needed
time = f.createDimension('time',None)
lat = f.createDimension('lat',4)
lon = f.createDimension('lon',4)

# create time and outfall variables
times = f.createVariable('time',np.float64,('time',))
lats = f.createVariable('latitude',np.float32,('lat',))
lons = f.createVariable('longitude',np.float32,('lon',))

# variables have shape (time,lat,lon)
flow = f.createVariable('flow',np.float32,('time','lat','lon'))
NO3 = f.createVariable('NO3',np.float32,('time','lat','lon'))
NH4 = f.createVariable('NH4',np.float32,('time','lat','lon'))
NO2 = f.createVariable('NO2',np.float32,('time','lat','lon'))
PO4 = f.createVariable('PO4',np.float32,('time','lat','lon'))
Fe = f.createVariable('Fe',np.float32,('time','lat','lon'))
SO2 = f.createVariable('SO2',np.float32,('time','lat','lon'))

BOD = f.createVariable('BOD',np.float32,('time','lat','lon'))
TOC = f.createVariable('TOC',np.float32,('time','lat','lon'))
ON = f.createVariable('ON',np.float32,('time','lat','lon'))
OP = f.createVariable('OP',np.float32,('time','lat','lon'))

alk = f.createVariable('alkalinity',np.float32,('time','lat','lon'))
pH = f.createVariable('pH',np.float32,('time','lat','lon'))
SiO4 = f.createVariable('sulfate',np.float32,('time','lat','lon'))
temp = f.createVariable('temperature',np.float32,('time','lat','lon'))
DO = f.createVariable('dissolved_oxygen',np.float32,('time','lat','lon'))
salinity = f.createVariable('salinity',np.float32,('time','lat','lon'))

flow.units = 'm3/s'
NO3.units = 'mmol/m3'
NH4.units =  'mmol/m3'
NO2.units ='mmol/m3'
PO4.units ='mmol/m3'
Fe.units = 'mmol/m3'
SO2.units = 'mmol/m3'

BOD.units = 'mmol/m3'
TOC.units = 'mmol/m3'
ON.units = 'mmol/m3'
OP.units = 'mmol/m3'

alk.units = 'mmol CaCO3/m3'
SiO4.units = 'mmol/m3'

temp.units = 'degrees C'
DO.units = 'mmol/m3'
salinity.units = 'psu'

lats.units = 'degrees north'
lons.units = 'degrees east'

times.units = 'days since 1970-12-31'

# set time dimension to be days since 1970-12-31
times[:] = date2num(potw_data_converted['PLWTP'][0],times.units)
lats[:] = lat_data
lons[:] = lon_data

# for potw_data_converted data is structured
# potw_data_converted[location][nutrient/flow]
# each key is one location
# 0 dimension for each key is date
# 1 dimension is flowrate m3/s
# 2 dimension is NO3 in mmol/m3
# 3 dimension is NO2 in mmol/m3
# 4 dimension is NH4 in mmol/m3
# 5              BOD    mmol/m3
# 6              COD    mmol/m3
# 7              Fe     mmol/m3
# 8              SO2    mmol/m3
# 9              ON     mmol/m3 
# 10             TP     mmol/m3
# 11             PO4    mmol/m3
# 12             OP     mmol/m3
# 13             TOC    mmol/m3

for t in range(len(times)):
    for l,p in enumerate(outfalls):
        flow[t,l,l] = potw_data_converted[p][1][t]
        NO3[t,l,l] = potw_data_converted[p][2][t]
        NO2[t,l,l] = potw_data_converted[p][3][t]
        NH4[t,l,l] = potw_data_converted[p][4][t]
        Fe[t,l,l] = potw_data_converted[p][7][t]
        SO2[t,l,l] = potw_data_converted[p][8][t]
        PO4[t,l,l] = potw_data_converted[p][11][t]

        BOD[t,l,l] = potw_data_converted[p][5][t]     
        TOC[t,l,l] = potw_data_converted[p][13][t]
        ON[t,l,l] = potw_data_converted[p][9][t]
        OP[t,l,l] = potw_data_converted[p][12][t]

        pH[t,l,l] = pH_data[l][t]
        alk[t,l,l] = alk_data_conv[l][t]
        SiO4[t,l,l] = SiO4_data_conv[l][t]

        temp[t,l,l] = temp_data[l][t]
        DO[t,l,l] = DO_data_conv[l][t]
        salinity[t,l,l] = salinity_data[l][t]

    print('assigning data for time '+str(t)+' of '+str(len(times)))

f.close()



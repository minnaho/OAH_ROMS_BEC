##################################################
# take in excel data from
# CMAQ_V5.0.1_2_2002_2012_TotalDep_kg_per_ha.xlsx
# and put into .nc file
# structure: time, lat, lon, no3, nh4, alkalinity, and iron
##################################################
import openpyxl
from openpyxl import Workbook
from netCDF4 import Dataset
import datetime
import time
import numpy as np
from netCDF4 import num2date, date2num

# load excel file
wb = openpyxl.load_workbook('CMAQ_v5.0.1_2_2002_2012_TotalDep_kg_per_ha_Minna_edits.xlsx',data_only=True)

# each sheet is a year in the excel file
model_year = wb.sheetnames

# make lists for each input to append data into
time_data = []
lat_data = np.zeros([113,299])
lon_data = np.zeros([113,299])

# oxidized nitrogen
oxn_data = []

# reduced nitrogen
redn_data = []

# total sulfur
s_data = []

# iron
fe_data = []

# append data
# lat/lon same for all years
# latitude and longitude both have two 2D array shape (113,299) because of curved grid
for i in range(2,33789):
    lon_data[wb[model_year[1]].cell(row=i,column=5).value-1,wb[model_year[1]].cell(row=i,column=6).value-1] = wb[model_year[1]].cell(row=i,column=1).value
    lat_data[wb[model_year[1]].cell(row=i,column=5).value-1,wb[model_year[1]].cell(row=i,column=6).value-1] = wb[model_year[1]].cell(row=i,column=2).value
    
     
# skip first worksheet because it does not have data
for year in model_year[1:]:  
    # make list of datetime objects for each month per year
    # e.g. 1/1/2002, 2/1/2002, ...
    for t in range(55,55+12):
        time_data.append(wb[year].cell(row=1,column=t).value)


# data is organized such that oxn_data[132n:132(n+1)] = data from oxn_Jan2002 to oxn_Dec2012 for one lat,lon coordinate
# where n is range(0,len(longitude coordinates))
for row_i in range(2,33789):
    for year in model_year[1:]:
        # j, h, k, l denotes column in excel sheet   
        for j in range(7,7+12): 
            if wb[year].cell(row=row_i,column=j).value == -999:
                oxn_data.append(np.nan)
            else:
                oxn_data.append(wb[year].cell(row=row_i,column=j).value)

        for h in range(19,19+12): 
            if wb[year].cell(row=row_i,column=h).value == -999:
                redn_data.append(np.nan)
            else:
                redn_data.append(wb[year].cell(row=row_i,column=h).value)

        for k in range(31,31+12): 
            if wb[year].cell(row=row_i,column=k).value == -999:
                s_data.append(np.nan)
            else:
                s_data.append(wb[year].cell(row=row_i,column=k).value)

        for l in range(43,43+12): 
            if wb[year].cell(row=row_i,column=l).value == -999:
                fe_data.append(np.nan)
            else:
                fe_data.append(wb[year].cell(row=row_i,column=l).value)
    print('appending raw data for row '+str(row_i)+' of 33789')
 
# convert data from kg/ha/month to mmol/m2/s
kg_to_mg = 1e6
N_mg_to_mmol = 1./14
S_mg_to_mmol = 1./32
Fe_mg_to_mmol = 1./56
hectare_to_m2 = 1./10000
month_to_seconds = 1./(86400*30)

oxn_data_conv = np.asarray(oxn_data) * N_mg_to_mmol * kg_to_mg * hectare_to_m2 * month_to_seconds

redn_data_conv = np.asarray(redn_data) * N_mg_to_mmol * kg_to_mg * hectare_to_m2 * month_to_seconds

s_data_conv = np.asarray(s_data) * S_mg_to_mmol * kg_to_mg * hectare_to_m2 * month_to_seconds

fe_data_conv = np.asarray(fe_data) * Fe_mg_to_mmol * kg_to_mg * hectare_to_m2 * month_to_seconds


# calculate alkalinity from reduced N, oxidized N, and sulfur
alk_data_conv = redn_data_conv - oxn_data_conv - (2*s_data_conv) 

# save data into .npy files to call later and save processing time
np.save('oxn_data_conv.npy',oxn_data_conv)
np.save('redn_data_conv.npy',redn_data_conv)
np.save('alk_data_conv.npy',alk_data_conv)
np.save('fe_data_conv.npy',fe_data_conv)
np.save('time_data.npy',time_data)
np.save('lon_data.npy',lon_data)
np.save('lat_data.npy',lat_data)

'''
#load data calculated previously to save time
time_data = np.load('time_data.npy')
lon_data = np.load('lon_data.npy')
lat_data = np.load('lat_data.npy')

oxn_data_conv = np.load('oxn_data_conv.npy')
redn_data_conv = np.load('redn_data_conv.npy')
fe_data_conv = np.load('fe_data_conv.npy')
alk_data_conv = np.load('alk_data_conv.npy')


# create netCDF file
f = Dataset('atmos_deposition_CMAQ_2002_2012.nc','w')

# details about data
f.title = 'Atmospheric Deposition from Reduced Nitrogen, Oxidized Nitrogen, Alkalinity, and Iron'
f.source = 'EPA Community Multiscale Air Quality modeling system (CMAQ) V5.0.2 monthly total deposition files 2002-2012 with adjusted wet deposition for continental US using 12km grids'
f.history = 'Created '+time.ctime(time.time())
f.comment = 'Alkalinity calculated by: Reduced N - Oxidized N - (2 * Total Sulfur)'

# create time, lat, lon dimensions
time = f.createDimension('time',None)
lat = f.createDimension('lat',lat_data.shape[0])
lon = f.createDimension('lon',lon_data.shape[1])

# create variables
times = f.createVariable('time',np.float64,('time',))
lats = f.createVariable('latitude',np.float32,('lat','lon',))
lons = f.createVariable('longitude',np.float32,('lat','lon',))

oxn = f.createVariable('oxidized nitrogen',np.float32,('time','lat','lon')) 
redn = f.createVariable('reduced nitrogen',np.float32,('time','lat','lon'))
alk = f.createVariable('alkalinity',np.float32,('time','lat','lon'))
fe = f.createVariable('iron',np.float32,('time','lat','lon'))

# units
times.units = 'days since 2002-1-1'
lats.units = 'degrees north'
lons.units = 'degrees east'
oxn.units = 'mmmol/m2/s'
redn.units = 'mmmol/m2/s'
alk.units = 'mmmol/m2/s'
fe.units = 'mmmol/m2/s'

times[:] = time_data[:]
lats[:,:] = lat_data[:,:]
lons[:,:] = lon_data[:,:]

# append data to nutrients
t = 0
for coord1 in range(lats.shape[1]):
    for coord2 in range(lats.shape[0]):
        oxn[:,coord2,coord1] = oxn_data_conv[len(times)*t:len(times)*(t+1)]
        redn[:,coord2,coord1] = redn_data_conv[len(times)*t:len(times)*(t+1)]
        fe[:,coord2,coord1] = fe_data_conv[len(times)*t:len(times)*(t+1)]
        alk[:,coord2,coord1] = alk_data_conv[len(times)*t:len(times)*(t+1)]
        t+=1
    print('assigned data for coordinate '+str(coord1)+' of '+str(lats.shape[1]))    

   
f.close()
print('data written to .nc file atmos_deposition_CMAQ_2002_2012.nc')
'''



